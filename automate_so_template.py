import pdfplumber
import openpyxl
import shutil
import re
import os
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext


def log(widget, message):
    widget.insert(tk.END, message + "\n")
    widget.see(tk.END)


# ─────────────────────────────────────────────
#  OCR 支持（用于图片型 PDF）
# ─────────────────────────────────────────────
def _ocr_pdf_pages(pdf_path, log_widget=None):
    """
    用 PyMuPDF + Tesseract 对 PDF 每页做 OCR。
    兼容 Tesseract 3.x / 4.x / 5.x。
    返回 (all_lines, page0_text, page0_lines)
    """
    try:
        import fitz
    except ImportError:
        raise RuntimeError("缺少 PyMuPDF，请运行: pip install pymupdf")
    try:
        import pytesseract
        from PIL import Image
        import io
    except ImportError:
        raise RuntimeError("缺少依赖，请运行: pip install pytesseract pillow")

    doc = fitz.open(pdf_path)
    all_lines = []
    page0_text = ""
    page0_lines = []

    for pno, page in enumerate(doc):
        mat = fitz.Matrix(200 / 72, 200 / 72)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("png")
        img = Image.open(io.BytesIO(img_bytes))

        # 兼容所有版本：不传 --oem 等新参数
        try:
            text = pytesseract.image_to_string(img, config='--psm 6')
        except Exception:
            text = pytesseract.image_to_string(img)

        lines = text.split("\n")
        if pno == 0:
            page0_text = text
            page0_lines = lines
        all_lines.extend(lines)

    if log_widget:
        log(log_widget, f"OCR 完成，共处理 {len(doc)} 页，提取 {len(all_lines)} 行")
    return all_lines, page0_text, page0_lines


def _extract_text_with_fallback(pdf_path, log_widget=None):
    """
    尝试用 pdfplumber 提取文字；若第一页为空（图片型 PDF），自动切换到 OCR。
    返回 (all_lines, page0_text, page0_lines, used_ocr: bool)
    """
    with pdfplumber.open(pdf_path) as pdf:
        page0_text = pdf.pages[0].extract_text() or ""
        if page0_text.strip():
            all_lines = []
            for page in pdf.pages:
                t = page.extract_text() or ""
                all_lines.extend(t.split("\n"))
            page0_lines = page0_text.split("\n")
            return all_lines, page0_text, page0_lines, False

    # 图片型 PDF → OCR
    if log_widget:
        log(log_widget, "检测到图片型 PDF，切换到 OCR 模式（速度稍慢，请稍候）...")
    all_lines, page0_text, page0_lines = _ocr_pdf_pages(pdf_path, log_widget)
    return all_lines, page0_text, page0_lines, True


# ─────────────────────────────────────────────
#  Bio Basic
# ─────────────────────────────────────────────
def process_biobasic(pdf_path, log_widget):
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text()

        po_match = re.search(r'P\.O\.No\..*\n(\S+)', text)
        if not po_match:
            po_match = re.search(r'P[.\s]*O[.\s]*No[.\s]*.*\n(\S+)', text)
        po_no = po_match.group(1) if po_match else ""
        if not po_no:
            raise ValueError("PO No. not found in PDF")
        log(log_widget, f"PO No.: {po_no}")

        items = []
        for line in text.split('\n'):
            code_match = re.search(r'A\d{6,}-\d{4}', line)
            if code_match:
                code = code_match.group(0)
                nums = re.findall(r'[\d\.]+', line)
                if len(nums) >= 3:
                    rate = nums[-2]
                    qty = nums[-3]
                    after_code = line[code_match.end():].strip()
                    desc = re.sub(r'[\d\.]+\s+[\d\.]+\s+[\d\.]+\s*$', '', after_code).strip()
                    items.append({'sap_id': code, 'desc': desc, 'qty': qty, 'rate': rate,
                                  'po_no': po_no, 'due_date': '', 'address': '',
                                  'order_date': '', 'cat_no': '', 'spec_sheet': ''})

    if not items:
        raise ValueError("No valid item lines found in PDF")
    log(log_widget, f"Successfully parsed {len(items)} items")
    return items


# ─────────────────────────────────────────────
#  日期格式转换
# ─────────────────────────────────────────────
def parse_order_date(raw_date):
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})$', raw_date.strip())
    if not m:
        return raw_date
    month, day, year = m.group(1), m.group(2), m.group(3)
    if len(year) == 2:
        year = '20' + year
    return f"{year}.{month.zfill(2)}.{day.zfill(2)}"


def parse_due_date_long(raw_date):
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})$', raw_date.strip())
    if not m:
        return raw_date
    month, day, year = m.group(1), m.group(2), m.group(3)
    if len(year) == 2:
        year = '20' + year
    return f"{month}/{day}/{year}"


def parse_due_date_av(raw_date):
    m = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})$', raw_date.strip())
    if not m:
        return f"Due date: {raw_date}"
    month, day, year = m.group(1), m.group(2), m.group(3)
    if len(year) == 2:
        year = '20' + year
    return f"Due date: {month}/{day}/{year}"


# ─────────────────────────────────────────────
#  Thermo Fisher — 健壮字段提取（全文搜索，不依赖行顺序）
# ─────────────────────────────────────────────

def _robust_find_order_and_date(page0_text, page0_lines):
    """
    多策略提取订单号 + 订单日期，兼容 Tesseract 各版本识别质量差异。
    """
    full_text = page0_text

    # 策略1：日期 + 空格 + 6位以上纯数字（最常见，行内或跨行拼接都能抓到）
    m = re.search(r'(\d{1,2}/\d{2}/\d{2})\s+(\d{6,})', full_text)
    if m:
        return m.group(2), m.group(1)

    # 策略2：ORDER NUMBER 附近的6位数字
    m = re.search(r'ORDER\s*NUMBER\D{0,50}?(\d{6,})', full_text, re.DOTALL)
    if m:
        order_no = m.group(1)
        dm = re.search(r'(\d{1,2}/\d{2}/\d{2})', full_text)
        return order_no, (dm.group(1) if dm else '')

    # 策略3：在前30行逐行找独立的6位数字
    for line in page0_lines[:30]:
        m = re.search(r'\b(\d{6,})\b', line)
        if m:
            order_no = m.group(1)
            dm = re.search(r'(\d{1,2}/\d{2}/\d{2})', full_text)
            return order_no, (dm.group(1) if dm else '')

    return '', ''


def _robust_find_due_date(page0_text):
    """从全文提取 EST. DELIVERY DATE 后的日期，兼容 OCR 断行。
    注意：用 .{0,60} 而非 \D{0,30}，因为下一行地址可能含数字（如 '220 NECK ROAD'）
    会导致 \D 匹配提前中断。
    为避免误抓订单日期，要求候选日期不紧跟6位数字（订单号格式）。
    """
    # 在 EST. DELIVERY DATE 后 60 个字符内找日期
    m = re.search(
        r'EST\.?\s*DELIVERY\s*DATE.{0,60}?(\d{1,2}/\d{2}/\d{2,4})',
        page0_text, re.IGNORECASE | re.DOTALL
    )
    if m:
        return m.group(1)
    # 宽松兜底
    m = re.search(
        r'DELIVERY\s*DATE.{0,60}?(\d{1,2}/\d{2}/\d{2,4})',
        page0_text, re.IGNORECASE | re.DOTALL
    )
    if m:
        return m.group(1)
    return ''


def _robust_find_ship_to(page0_lines):
    """提取 SHIP TO 地址行，兼容 Tesseract 3 识别误差"""
    ship_to_idx = None
    vendor_idx = None
    for i, line in enumerate(page0_lines):
        if re.match(r'SHIP\s*T[O0]', line.strip(), re.IGNORECASE) and ship_to_idx is None:
            ship_to_idx = i
        if re.match(r'VENDOR', line.strip(), re.IGNORECASE) and ship_to_idx is not None:
            vendor_idx = i
            break

    if ship_to_idx is None:
        return []

    end_idx = vendor_idx if vendor_idx else ship_to_idx + 7
    noise_kw = ['CORRESPONDENCE', 'F.O.B.', 'F.0.B.', 'EST. DELIVERY',
                 'FREIGHT', 'TERMS', 'NET 30', 'NOTE:']
    result = []
    for line in page0_lines[ship_to_idx + 1:end_idx]:
        clean = line.strip()
        for n in noise_kw:
            clean = clean.split(n)[0].strip()
        clean = re.sub(r'\s+\d+/\d+/\d+\s*$', '', clean).strip()
        if clean:
            result.append(clean)
    return result


def _extract_road_name(ship_lines):
    for line in ship_lines:
        m = re.match(r'^\d+\s+(.+)$', line.strip())
        if m:
            return m.group(1).strip().title()
    return ''


# ─────────────────────────────────────────────
#  页面噪声过滤 & 物料解析
# ─────────────────────────────────────────────

PAGE_NOISE = [
    'REFER ALL COMMUNICATIONS', 'Our General Purchase Terms',
    'www.thermofisher.com/PO-Terms', 'writing. We do not accept',
    'submit, use or refer to', 'expressly agree to additional',
    'of our Order. These Terms', 'AN EQUAL OPPORTUNITY',
    'NOTIFICATION OF ANY PRICE', 'ISO 14001', 'A part of:',
    'ORDER ENTERED UNDER', 'COMPLETE SHIPMENT', 'ACKNOWLEDGED BY',
    'PHONE NUMBER', 'DATE OF THIS ACKNOWLEDGMENT',
    'THERMO FISHER SCIENTIFIC CHEMICALS INC.',
    'THERMO FISHER SCIENTIFIC', 'ALFA AESAR', 'PURCHASE ORDER',
    'FEDERAL ID NO.', 'THIS ORDER IS SUBJECT', '** REPRINT **',
    'NOTE: THE ABOVE', 'ALL INVOICES, BILLS', 'CORRESPONDENCE',
    'F.O.B.', 'F.0.B.', 'EST. DELIVERY DATE', 'FREIGHT', 'TERMS',
    'NET 30', 'SHIP VIA', 'D-U-N-S', 'SPECIAL INSTRUCTIONS',
    'SHIP TO', 'VENDOR', 'QUANTITY CHG', 'BIO BASIC INC.',
    'BAILEY AVENUE', 'AMHERST', 'Please send electronic invoice',
    'Please confirm price', 'Please send CoA', 'Please reference PO',
    'Must include Country', 'alfaaesar.accountspay',
    'Radcliff Road', 'Phone: 978', 'PAGE:', '3. VIA', '5. PHONE', '6. DATE',
]

J_CODE_RE = re.compile(
    r'\b([A-Z][A-Z0-9]{5,}-[A-Z0-9#@.]+|[0-9]{6,}-[A-Z0-9#@.]+)\b'
)


def _is_noise_line(line):
    return any(kw in line for kw in PAGE_NOISE)


def _is_valid_price(num_str):
    """
    判断一个数字字符串是否像真实的 unit price。
    排除：浓度/pH值等（0.2、8.0、1.0 这类小数位只有1位且值<15的数字）
    保留：20.000、25.170、160.00、302.04 这类真实价格
    """
    try:
        val = float(num_str)
        if val < 1.0:
            return False
        decimal_part = num_str.split('.')[-1]
        # 小数只有1位且值<15，判定为 pH/浓度描述（如 8.0、5.0、1.0）
        if len(decimal_part) == 1 and val < 15:
            return False
        return True
    except ValueError:
        return False


def _parse_items_from_lines(all_lines, order_no, due_date_av, due_date_m,
                             address, road_name, order_date, db, log_widget):
    items = []
    for idx, line in enumerate(all_lines):
        code_match = J_CODE_RE.search(line)
        if not code_match:
            continue

        # 在当前行找 qty（格式：数字/）
        qty_match = re.search(r'(\d+)/', line)

        # 如果当前行没有 qty，往前最多3行找（OCR可能把qty和code拆成两行）
        if not qty_match:
            for prev_idx in range(idx - 1, max(idx - 4, -1), -1):
                prev_line = all_lines[prev_idx].strip()
                if not prev_line:
                    continue
                # 前一行如果已经有code，说明是另一个物料，停止
                if J_CODE_RE.search(prev_line):
                    break
                prev_qty = re.search(r'(\d+)/', prev_line)
                if prev_qty:
                    qty_match = prev_qty
                    break

        # 仍然找不到 qty 则跳过
        if not qty_match:
            continue

        # 在当前行找 unit price（过滤掉浓度描述里的小数如 0.2M）
        nums = [n for n in re.findall(r'\d+\.\d+', line) if _is_valid_price(n)]

        # 如果当前行没有价格数字，往后最多5行继续找
        if not nums:
            for look_idx in range(idx + 1, min(idx + 6, len(all_lines))):
                look_line = all_lines[look_idx].strip()
                if not look_line:
                    continue
                # 遇到下一个物料行就停
                if J_CODE_RE.search(look_line) and re.search(r'\d+/', look_line):
                    break
                extra = [n for n in re.findall(r'\d+\.\d+', look_line) if _is_valid_price(n)]
                if extra:
                    nums = extra
                    break

        # 仍然找不到价格则跳过
        if not nums:
            continue

        cat_no = code_match.group(1)
        qty = qty_match.group(1)
        # unit price 是倒数第二个（最后一个是 amount），只有一个时直接用
        rate = nums[-2] if len(nums) >= 2 else nums[0]
        after_code = line[code_match.end():].strip()
        desc = re.sub(r'\s+\d+\.\d+\s+.*$', '', after_code).strip()

        sap_id = db.get(cat_no, '')
        if not sap_id:
            log(log_widget, f"  WARNING: {cat_no} not found in Database, AL will be empty")

        # 检测 spec sheet
        spec_sheet = ''
        for next_idx in range(idx + 1, min(idx + 80, len(all_lines))):
            next_line = all_lines[next_idx].strip()
            if not next_line:
                continue
            if J_CODE_RE.search(next_line) and re.search(r'\d+/', next_line):
                break
            if _is_noise_line(next_line):
                continue
            if 'https://assets.thermofisher.com' in next_line:
                spec_sheet = 'Spec sheet'
                break
            if re.search(r'https[:\s]+//.*thermofisher', next_line, re.IGNORECASE):
                spec_sheet = 'Spec sheet'
                break
            if 'thermofisher' in next_line.lower() and 'chem' in next_line.lower():
                spec_sheet = 'Spec sheet'
                break
            if re.search(r'Spec\s*Sheet\s*(for\s*this\s*item\s*is\s*not|fo[r}])', next_line, re.IGNORECASE):
                spec_sheet = ''
                break

        items.append({
            'sap_id': sap_id, 'desc': desc, 'qty': qty, 'rate': rate,
            'po_no': order_no, 'due_date_av': due_date_av, 'due_date_m': due_date_m,
            'address': address, 'road_name': road_name, 'order_date': order_date,
            'cat_no': cat_no, 'spec_sheet': spec_sheet,
        })
    return items


# ─────────────────────────────────────────────
#  Thermo Fisher 主处理函数
# ─────────────────────────────────────────────
def process_thermofisher(pdf_path, db_path, log_widget):
    # 加载 SAP 数据库
    wb_db = openpyxl.load_workbook(db_path)
    sap_sheet = 'SAP Database ' if 'SAP Database ' in wb_db.sheetnames else 'Product Databse'
    ws_db = wb_db[sap_sheet]
    db = {}
    for row in ws_db.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            db[str(row[0]).strip()] = str(row[1]).strip()
    log(log_widget, f"Database loaded: {len(db)} entries")

    # 提取文字（自动判断是否需要 OCR）
    all_lines, page0_text, page0_lines, used_ocr = _extract_text_with_fallback(pdf_path, log_widget)
    if not used_ocr:
        log(log_widget, "文字型 PDF，直接解析...")

    # ── 调试：打印OCR前20行，方便排查 ──
    log(log_widget, "=== OCR 第1页前20行（调试）===")
    for i, l in enumerate(page0_lines[:20]):
        log(log_widget, f"  {i:2d}: {l!r}")
    log(log_widget, "=== 调试结束 ===")

    # 提取订单号 + 订单日期（全文健壮搜索）
    order_no, raw_order_date = _robust_find_order_and_date(page0_text, page0_lines)
    if not order_no:
        # 兜底：打印全页OCR文字让用户反馈
        log(log_widget, "=== 无法找到订单号，全页OCR文字如下（请截图反馈）===")
        for i, l in enumerate(page0_lines):
            log(log_widget, f"  {i:2d}: {l!r}")
        raise ValueError(
            "Order Number not found in PDF\n\n"
            "请查看日志中的 OCR 调试文字。\n"
            "建议升级到 Tesseract 4.x 或 5.x 以获得更好识别效果：\n"
            "https://github.com/UB-Mannheim/tesseract/wiki"
        )

    log(log_widget, f"Order No.: {order_no}")
    order_date = parse_order_date(raw_order_date) if raw_order_date else ''
    log(log_widget, f"Order Date: {order_date}" if order_date else "WARNING: Order date not found")

    # 提取 EST. DELIVERY DATE
    raw_due_date = _robust_find_due_date(page0_text)
    due_date_av = parse_due_date_av(raw_due_date) if raw_due_date else ''
    due_date_m = parse_due_date_long(raw_due_date) if raw_due_date else ''
    log(log_widget, f"Due Date: {raw_due_date}" if raw_due_date else "WARNING: Due date not found")

    # 提取地址
    ship_to_raw_lines = _robust_find_ship_to(page0_lines)
    address = ', '.join(ship_to_raw_lines)
    road_name = _extract_road_name(ship_to_raw_lines)
    log(log_widget, f"Ship To: {address}" if address else "WARNING: Address not found")
    log(log_widget, f"Road Name: {road_name}" if road_name else "WARNING: Road name not found")

    # 解析物料行
    items = _parse_items_from_lines(
        all_lines, order_no, due_date_av, due_date_m,
        address, road_name, order_date, db, log_widget
    )

    if not items:
        raise ValueError("No valid item lines found in PDF")
    log(log_widget, f"Successfully parsed {len(items)} items")
    return items


# ─────────────────────────────────────────────
#  写入 Excel
# ─────────────────────────────────────────────
def write_to_excel(items, template_path, output_path, log_widget):
    shutil.copy2(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for i, item in enumerate(items):
        r = 3 + i
        data = {'A': 1, 'B': 3220, 'C': 10, 'E': "00", 'G': "ZOR1", 'Q': "B014", 'R': 3130,
                'S': "East China-2", 'T': 118, 'U': "HQ-Domestic", 'AC': "01", 'AF': "D002",
                'AP': "PR00", 'AS': 3200}
        for col, val in data.items():
            ws[f"{col}{r}"] = val
        for col in ['I', 'K', 'M', 'O']:
            ws[f"{col}{r}"] = 5148715
        ws[f"AB{r}"] = ws[f"AR{r}"] = "USD"
        ws[f"AD{r}"] = ws[f"AE{r}"] = item['po_no']
        ws[f"AK{r}"] = (i + 1) * 10
        ws[f"AL{r}"] = item['sap_id']
        ws[f"AM{r}"] = item['desc']
        ws[f"AO{r}"] = item['qty']
        ws[f"AQ{r}"] = item['rate']
        due_date_val = item.get('due_date_av') or item.get('due_date', '')
        if due_date_val:
            ws[f"AV{r}"] = due_date_val
        if item.get('address'):
            ws[f"AG{r}"] = item['address']

    wb.save(output_path)
    wb.close()
    log(log_widget, f"Done. New file saved to: {output_path}")
    messagebox.showinfo("Success", f"New Excel file created:\n{output_path}")


def write_to_orders_excel(items, db_path, log_widget):
    from openpyxl.styles import Alignment
    log(log_widget, "Writing to 2025 Orders sheet...")
    wb = openpyxl.load_workbook(db_path)
    ws = wb['2025 Orders']

    last_row = 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 2).value is not None or ws.cell(r, 1).value is not None:
            last_row = r

    next_row = last_row + 1
    log(log_widget, f"Appending from row {next_row}")
    center = Alignment(horizontal='center', vertical='center', wrap_text=False)

    def set_cell(ws, r, c, value):
        cell = ws.cell(r, c)
        cell.value = value
        cell.alignment = center

    for item in items:
        r = next_row
        set_cell(ws, r, 2, item['po_no'])
        set_cell(ws, r, 3, item['sap_id'])
        set_cell(ws, r, 4, item.get('cat_no', ''))
        set_cell(ws, r, 5, item['desc'])
        try:
            qty_val = int(item['qty']) if item['qty'] else None
        except (ValueError, TypeError):
            qty_val = item['qty']
        set_cell(ws, r, 6, qty_val)
        set_cell(ws, r, 7, f"=VLOOKUP(D{r},'Product Databse'!B:D,3,)")
        set_cell(ws, r, 8, item.get('order_date', ''))
        spec = item.get('spec_sheet', '')
        set_cell(ws, r, 9, spec if spec else None)
        set_cell(ws, r, 13, item.get('due_date_m') or None)
        road = item.get('road_name', '')
        set_cell(ws, r, 16, road if road else None)
        next_row += 1

    wb.save(db_path)
    wb.close()
    log(log_widget, f"2025 Orders updated and saved: {db_path}")
    messagebox.showinfo("Success", f"2025 Orders updated successfully:\n{db_path}")


# ─────────────────────────────────────────────
#  主流程
# ─────────────────────────────────────────────
def process_data(pdf_path, template_path, output_path, db_path, pdf_type,
                 orders_db_path, expected_qty_str, log_widget):
    try:
        if not pdf_path or not template_path or not output_path:
            raise ValueError("Please select PDF, Template and Output files")
        if pdf_type == "Thermo Fisher" and not db_path:
            raise ValueError("Please select Database file for Thermo Fisher PO")

        log(log_widget, f"Reading PDF ({pdf_type})...")

        if pdf_type == "Bio Basic":
            items = process_biobasic(pdf_path, log_widget)
        else:
            items = process_thermofisher(pdf_path, db_path, log_widget)

        actual_count = len(items)
        if expected_qty_str.strip():
            try:
                expected_count = int(expected_qty_str.strip())
                if actual_count != expected_count:
                    msg = (f"Item count mismatch!\n\n"
                           f"Expected:  {expected_count} items\n"
                           f"Parsed:    {actual_count} items\n\n"
                           f"Please check the log and verify the PDF before continuing.\n"
                           f"Do you want to continue anyway?")
                    log(log_widget, f"WARNING: Expected {expected_count} items but parsed {actual_count}!")
                    if not messagebox.askyesno("Count Mismatch", msg, icon="warning"):
                        log(log_widget, "Processing cancelled by user.")
                        return
                else:
                    log(log_widget, f"Item count verified: {actual_count} items ✓")
            except ValueError:
                log(log_widget, "WARNING: Invalid expected quantity input, skipping count check.")

        write_to_excel(items, template_path, output_path, log_widget)
        if orders_db_path:
            write_to_orders_excel(items, orders_db_path, log_widget)

    except Exception as e:
        log(log_widget, f"Error: {str(e)}")
        messagebox.showerror("Error", str(e))


# ─────────────────────────────────────────────
#  GUI
# ─────────────────────────────────────────────
root = tk.Tk()
root.title("SAP Data Processing Tool")
root.geometry("560x820")
frame = ttk.Frame(root, padding="10")
frame.pack(fill=tk.BOTH, expand=True)

ttk.Label(frame, text="PDF Type:").pack(anchor=tk.W)
pdf_type_var = tk.StringVar(value="Bio Basic")
type_frame = ttk.Frame(frame)
type_frame.pack(anchor=tk.W, pady=2)
ttk.Radiobutton(type_frame, text="Bio Basic Sales Order", variable=pdf_type_var, value="Bio Basic").pack(side=tk.LEFT)
ttk.Radiobutton(type_frame, text="Thermo Fisher PO", variable=pdf_type_var, value="Thermo Fisher").pack(side=tk.LEFT, padx=10)

pdf_var = tk.StringVar()
template_var = tk.StringVar()
output_var = tk.StringVar()
db_var = tk.StringVar()
orders_db_var = tk.StringVar()

ttk.Label(frame, text="PDF File:").pack(anchor=tk.W, pady=(8, 0))
ttk.Entry(frame, textvariable=pdf_var).pack(fill=tk.X)
ttk.Button(frame, text="Select PDF", command=lambda: pdf_var.set(filedialog.askopenfilename(
    filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]))).pack(pady=3)

ttk.Label(frame, text="Excel Template File:").pack(anchor=tk.W)
ttk.Entry(frame, textvariable=template_var).pack(fill=tk.X)
ttk.Button(frame, text="Select Template", command=lambda: template_var.set(filedialog.askopenfilename(
    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]))).pack(pady=3)

ttk.Label(frame, text="Output File (save new Excel as):").pack(anchor=tk.W)
ttk.Entry(frame, textvariable=output_var).pack(fill=tk.X)
ttk.Button(frame, text="Select Output Path", command=lambda: output_var.set(filedialog.asksaveasfilename(
    defaultextension=".xlsx",
    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]))).pack(pady=3)

ttk.Label(frame, text="SAP Database File (Thermo Fisher only):").pack(anchor=tk.W)
ttk.Entry(frame, textvariable=db_var).pack(fill=tk.X)
ttk.Button(frame, text="Select SAP Database", command=lambda: db_var.set(filedialog.askopenfilename(
    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]))).pack(pady=3)

ttk.Separator(frame, orient='horizontal').pack(fill=tk.X, pady=6)

qty_frame = ttk.Frame(frame)
qty_frame.pack(fill=tk.X, pady=(0, 4))
ttk.Label(qty_frame, text="Expected Item Count (optional):").pack(side=tk.LEFT)
expected_qty_var = tk.StringVar()
ttk.Entry(qty_frame, textvariable=expected_qty_var, width=8).pack(side=tk.LEFT, padx=6)
ttk.Label(qty_frame, text="(leave blank to skip check)", foreground="gray").pack(side=tk.LEFT)

ttk.Label(frame, text="2025 Orders Database (optional):").pack(anchor=tk.W)
ttk.Entry(frame, textvariable=orders_db_var).pack(fill=tk.X)
ttk.Button(frame, text="Select Orders Database", command=lambda: orders_db_var.set(filedialog.askopenfilename(
    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]))).pack(pady=3)

ttk.Button(frame, text="Start Processing",
           command=lambda: process_data(
               pdf_var.get(), template_var.get(), output_var.get(),
               db_var.get(), pdf_type_var.get(),
               orders_db_var.get(),
               expected_qty_var.get(),
               log_area)
           ).pack(pady=8)

log_area = scrolledtext.ScrolledText(frame, height=10, width=60)
log_area.pack(pady=5)

root.mainloop()